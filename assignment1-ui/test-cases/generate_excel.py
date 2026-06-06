"""Generate SauceDemo manual test cases as a formatted Excel workbook."""
import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

CSV_PATH = Path(__file__).parent / "SauceDemo_Manual_TestCases.csv"
XLSX_PATH = Path(__file__).parent / "SauceDemo_Manual_TestCases.xlsx"

# Mandatory columns from the assignment (plus helpful extras).
HEADERS = [
    "Test Case ID",
    "Module",
    "Scenario Type",
    "Title",
    "Preconditions",
    "Severity",
    "Priority",
    "Steps",
    "Expected",
    "Actual",
    "Status",
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")


def load_rows():
    with CSV_PATH.open(encoding="utf-8", newline="") as fh:
        reader = csv.DictReader(fh)
        return list(reader)


def write_sheet(ws, rows):
    ws.append(HEADERS)
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        ws.append([row.get(h, "") for h in HEADERS])

    from openpyxl.utils import get_column_letter

    widths = [14, 12, 14, 38, 30, 10, 10, 50, 45, 12, 10]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = WRAP

    ws.freeze_panes = "A2"


def main():
    rows = load_rows()
    wb = Workbook()

    # Sheet 1: All test cases
    all_sheet = wb.active
    all_sheet.title = "All Test Cases"
    write_sheet(all_sheet, rows)

    # Sheet 2/3/4: filtered by scenario type
    for scenario, title in [
        ("Positive", "Positive Test Cases"),
        ("Negative", "Negative Test Cases"),
        ("E2E", "End to End Test Cases"),
    ]:
        ws = wb.create_sheet(title)
        filtered = [r for r in rows if r["Scenario Type"] == scenario]
        write_sheet(ws, filtered)

    wb.save(XLSX_PATH)
    print(f"Created: {XLSX_PATH}")
    print(f"  All:      {len(rows)} cases")
    print(f"  Positive: {sum(1 for r in rows if r['Scenario Type'] == 'Positive')}")
    print(f"  Negative: {sum(1 for r in rows if r['Scenario Type'] == 'Negative')}")
    print(f"  E2E:      {sum(1 for r in rows if r['Scenario Type'] == 'E2E')}")


if __name__ == "__main__":
    main()
